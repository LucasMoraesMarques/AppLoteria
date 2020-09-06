from openpyxl import load_workbook

book = load_workbook('Jogo.xlsx')

conc = int(input('Digite o concurso atual:'))
data = str(input('Digite a data:'))
print(f'Digite o resultado do concurso nº{conc} do dia {data}:')
result = set(int(input(f'Dígito {i + 1}: ')) for i in range(0, 15))
nome_jogos = book.sheetnames


def acertos(result, jogos):
    x = 0
    acertos = []
    for k, v in jogos.items():
        x = len(v & result)
        acertos.append([k, x])
    return acertos


file = 'resultados.txt'
f = open(file, 'wt')
f.write(f'Resultado referente ao Concurso nº{conc}({data}) da Lotofácil\n')
for nome in nome_jogos:
    sheet = book[f'{nome}']
    l = sheet.max_row
    jogos = dict()
    for row in sheet.iter_rows(min_row=1, max_row=l, max_col=16):
        for i, cell in enumerate(row):
            if i == 0:
                name = cell.value
                jogos[name] = set()
            else:
                jogos[name].add(int(cell.value))
    acerto = acertos(result, jogos)
    print(f'\n{"-=" * 5} {nome} {"-=" * 5}\n')
    f.write(f'\n{"-=" * 5} {nome} {"-=" * 5}\n')
    for j_name, pts in acerto:
        print(f'{j_name}: {pts} acertos')
        f.write(f'{j_name}: {pts} acertos\n')

f.close()

"""#for i in range(1, l+1):
#   for j in range(1, 17):
        #print(sheet.cell(row=i, column=j).value)
        k = sheet.cell(row=i, column=j).value
        if k == f'Jogo {i}':
            jogos[k] = set()
        else:
            jogos[f'Jogo {i}'].add(int(k))
            
acerto = acertos(result, jogos)
print(f'\n{"-="*5} {nome} {"-="*5}\n')
f.write(f'\n{"-="*5} {nome} {"-="*5}\n')
for i in range(1, l+1):
    print(f'Jogo {i}: {acerto[i-1]} acertos')
    f.write(f'Jogo {i}: {acerto[i-1]} acertos\n')"""

# f.close()
