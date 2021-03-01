import numpy as np
import pandas as pd
import csv
from ProjetosLoteria.App.Sorteio import Sorteio, Loterias, Jogos
import requests
from openpyxl import load_workbook
import pprint


url = 'https://cutt.ly/LfnJrGv'


def carrega_resultados():
    resultados = pd.read_excel('resultados.xlsx', index_col=0, parse_dates=['Data'])
    cAtual = resultados.index[0]
    loto = Loterias('Lotofácil', cAtual, resultados)
    return loto


def busca_resultado(url):
    while True:
        resp = requests.get(url)
        if resp.status_code == 200:
            resultado = resp.json()
            break
        else:
            continue
    return resultado


def atualiza_resultados(dtf, resultado, loto):
    if float(resultado['numero']) != loto.cAtual:
        res = pd.DataFrame([[resultado['dataApuracao']] + resultado['listaDezenas']], columns=dtf.columns, index=[int(float(resultado["numero"]))])
        dt = dtf.append(res)
        dt.sort_index(inplace=True, ascending=False, axis=0)
        dt.to_excel('resultados.xlsx')
        return dt
    else:
        print('O resultado do concurso atual ainda não está disponível')


def confere_jogos(Loto, conc):
    book = load_workbook('Jogo.xlsx')
    nome_jogos = book.sheetnames
    file = 'resultados.txt'
    f = open(file, 'wt')
    f.write(f'Resultado referente ao Concurso nº{Loto.cAtual}({Loto.resultados.iloc[0, 0]}) da Lotofácil\n')
    for nome in nome_jogos:
        sheet = book[f'{nome}']
        l = sheet.max_row
        jogos = dict()
        for row in sheet.iter_rows(min_row=1, max_row=l, max_col=16):
            for i, cell in enumerate(row):
                if i == 0:
                    name = cell.value
                    jogos[name] = set()
                else:
                    jogos[name].add(int(cell.value))
        jogos = Jogos(conc, jogos)
        acertos = jogos.confere_acertos()
        f.write(f'\n{"-=" * 5} {nome} {"-=" * 5}\n')
        #print(f'\n{"-=" * 5} {nome} {"-=" * 5}\n')
        for j_name, pts in acertos:
        #    print(f'{j_name}: {pts} acertos')
            f.write(f'{j_name}: {pts} acertos\n')

    f.close()


dt = pd.read_excel('resultados.xlsx', index_col=0, header=0)

dt = atualiza_resultados(dt, resultado, Loto)
print(Loto.res_atual)
print(dt)
dt.to_csv("dt_resultados.csv")





